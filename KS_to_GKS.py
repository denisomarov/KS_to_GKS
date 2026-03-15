import streamlit as st
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, Alignment

def process_with_openpyxl(uploaded_file, sheet_name):

    # Инииализация переменных

    PSName                  = 'Туркибас'
    sign_of_existing_cable  = 'Существующий кабель'
    opt_c_name              = ['6XV8100 LC-LC', '6XV8100 ST-ST', '6XV8100 ST-LC', '6XV8100 LC-ST']

    # Инициализация структуры для сохранения результата
    data           = {'Тип кабеля': [], 'Код заказа': [], 'Завод-изготовитель': [], 'Длина, м': [], 'Кол-во, шт': []}
    df_cable       = pd.DataFrame(data)

    # Подготовка таблицы к работе

    df_a = pd.read_excel(uploaded_file, sheet_name=sheet_name)

    if not df_a.empty:

        df_a = df_a.iloc[1:,1:]
        df_a.columns = ['№ кабеля', 'Марка кабеля', 'Жильность x сечение',
               'Кол-во использ. Жил', 'Откуда', 'Куда',
               'Длина проект, м', 'Длина факт, м', 'Примечание']

        # Очистка полученных данных и формирование базы
        # Убираем непечатные символы
        df_a = df_a.apply(lambda x: x.str.replace(r'[\r\n\t]', '', regex=True) if x.dtype == "str" else x)

        df_a['Откуда'] = df_a['Откуда'].astype(str)
        df_a['Куда'] = df_a['Куда'].astype(str)
        df_a['Длина проект, м'] = df_a['Длина проект, м'].astype(str)
        df_a['Длина факт, м'] = df_a['Длина факт, м'].astype(str)

        # Убираем существующий кабель
        df_a = df_a.loc[df_a['Примечание'] != sign_of_existing_cable]

        # выделяем кабель для ГСИКБ

        for patch_name in opt_c_name:

            df_optic = df_a.loc[df_a['Марка кабеля'] == patch_name]
            df_optic['Марка кабеля'] = df_optic['Марка кабеля'] + ' ' + df_optic['Длина проект, м'] + ' м'

            df_optic_sorted = df_optic.groupby(['Марка кабеля', 'Длина проект, м']).agg('count').iloc[:, 0:1]
            df_optic_sorted.reset_index(inplace=True)
            df_optic_sorted.columns = ['Тип кабеля', 'Длина, м', 'Кол-во, шт']
            df_optic_sorted['Длина, м'] = df_optic_sorted['Длина, м'].astype(float)
            df_optic_sorted['Кол-во, шт'] = df_optic_sorted['Кол-во, шт'].astype(int)
            df_optic_sorted.sort_values(['Длина, м'], inplace=True)

            df_optic_sorted['Код заказа'] = ''
            df_optic_sorted['Завод-изготовитель'] = 'SIEMENS'
            df_optic_sorted = df_optic_sorted[['Тип кабеля', 'Код заказа', 'Завод-изготовитель', 'Длина, м', 'Кол-во, шт']]

            if len(df_optic_sorted) > 1:
                df_cable = pd.concat([df_cable, df_optic_sorted], ignore_index=True)

        # фиксируем формат финальной таблицы ГСИКБ
        df_cable['Тип кабеля'] = df_cable['Тип кабеля'].astype(str)
        df_cable['Код заказа'] = df_cable['Код заказа'].astype(str)
        df_cable['Завод-изготовитель'] = df_cable['Завод-изготовитель'].astype(str)
        df_cable['Длина, м'] = df_cable['Длина, м'].astype(float)
        df_cable['Кол-во, шт'] = df_cable['Кол-во, шт'].astype(int)

        # сохраняем оптический кабель в таблице для сметы
        df_cable_smeta = df_cable.copy()

        # обработка медного кабеля

        # выделяем медный кабель
        df_cupper = df_a.loc[~df_a['Марка кабеля'].isin(opt_c_name)]
        df_cupper['Марка кабеля'] = df_cupper['Марка кабеля'] + ' ' + df_cupper['Жильность x сечение']

        # выделяем медный кабель для сметы
        df_cupper_smeta = df_cupper.copy()

        # обработка медного кабеля общего (включая внутри шкафов)
        df_cupper = df_cupper[['Марка кабеля', 'Длина проект, м']]
        df_cupper.columns = ['Тип кабеля', 'Длина, м']
        df_cupper['Длина, м'] = df_cupper['Длина, м'].astype(float)
        df_cupper = df_cupper.groupby('Тип кабеля').agg('sum')
        df_cupper.sort_index()

        # оформляем таблицу медного кабеля общего (включая внутри шкафов)
        df_cupper.reset_index(inplace=True)
        df_cupper['Код заказа'] = ''
        df_cupper['Завод-изготовитель'] = ''
        df_cupper['Кол-во, шт'] = 1
        df_cupper = df_cupper[['Тип кабеля', 'Код заказа', 'Завод-изготовитель', 'Длина, м', 'Кол-во, шт']]

        # заполняем полную таблицу кабеля
        if len(df_cupper) > 1:
            df_cable = pd.concat([df_cable, df_cupper], axis=0, ignore_index=True)

        # выгружаем Групповую спецификацию изделий кабельной продукции в файл Excel
        wb = Workbook()
        ws = wb.active

        # заполняем ячейки из таблицы
        for r in dataframe_to_rows(df_cable, index=True, header=True):
            ws.append(r)

        # зададим ширину столбцов 20 единиц
        ws.column_dimensions['A'].width  = 5
        ws.column_dimensions['B'].width  = 20
        ws.column_dimensions['C'].width  = 20
        ws.column_dimensions['D'].width  = 20
        ws.column_dimensions['E'].width  = 15
        ws.column_dimensions['F'].width  = 15

        # зададим стиль первого ряда - заголовка таблицы
        # Создание стиля шрифта - жирный
        bold_font = Font(bold=True, color="000000") # Жирный, черный

        # Создание стиля выравнивания ячеек
        center_center = Alignment(horizontal='center', vertical='center') # Центрирование значения ячейки
        right_center  = Alignment(horizontal='right', vertical='center') # Выравнивание по горизонтали - правое, во вертикали - центр

        # Создание стиля границы ячейки
        thin_side = Side(style='thin', color="000000")  # Черная тонкая граница
        cell_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side) # Черная тонкая граница по всему периметру ячейки

        # применение стиля к столбцам

        for cell in ws['A']:
            cell.alignment = center_center
            cell.border = cell_border

        for cell in ws['B']:
            cell.alignment = center_center
            cell.border = cell_border

        for cell in ws['C']:
            cell.alignment = center_center
            cell.border = cell_border

        for cell in ws['D']:
            cell.alignment = center_center
            cell.border = cell_border

        for cell in ws['E']:
            cell.alignment = right_center
            cell.border = cell_border
            cell.number_format = '### ### ##0.00' # формат численного значения ячейки

        for cell in ws['F']:
            cell.alignment = right_center
            cell.border = cell_border

        # применение стилей к ячейкам заголовка

        col_A1 = ws['A1']
        col_A1.font = bold_font
        col_A1.alignment = center_center

        cell_B1 = ws['B1']
        cell_B1.font = bold_font
        cell_B1.alignment = center_center

        cell_C1 = ws['C1']
        cell_C1.font = bold_font
        cell_C1.alignment = center_center

        cell_D1 = ws['D1']
        cell_D1.font = bold_font
        cell_D1.alignment = center_center

        cell_E1 = ws['E1']
        cell_E1.font = bold_font
        cell_E1.alignment = center_center

        cell_F1 = ws['F1']
        cell_F1.font = bold_font
        cell_F1.alignment = center_center

        # запись результата в буфер
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


# --- Интерфейс Streamlit ---
st.title("Групповая кабельная спецификация")

uploaded_file = st.file_uploader("Загрузите Excel", type=['xlsx'])

if uploaded_file is not None:
    wb_temp = load_workbook(uploaded_file, read_only=True)
    sheets = wb_temp.sheetnames
    selected_sheet = st.selectbox("Выберите лист:", sheets)

    if st.button("✨ Обработать файл"):
        # Важно: для openpyxl нужно сбросить указатель файла
        uploaded_file.seek(0)
        result = process_with_openpyxl(uploaded_file, selected_sheet)

        st.download_button(
            label="📥 Скачать готовый файл",
            data=result,
            file_name="formatted_data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
